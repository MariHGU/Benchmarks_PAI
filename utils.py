from pathlib import Path
import logging, os, time
from datetime import datetime
from hashlib import sha1
from enum import IntEnum
from typing import Tuple, List
from ollama import AsyncClient
from openpyxl import load_workbook
import pandas as pd

class TypeOfTest(IntEnum):
    SUMMARIZATION = 1
    PROMPT_ALIGNMENT = 2
    HELPFULNESS = 3
    BENCHMARKING = 4

class CustomRelativeFormatter(logging.Formatter):
    """
    ChatGPT-4o generated code:
    A custom logging formatter that formats the time of the log record
    """

    def formatTime(self, record, datefmt=None):
        # relativeCreated is in milliseconds
        total_seconds = record.relativeCreated / 1000
        minutes = int(total_seconds // 60)
        seconds = total_seconds % 60
        return f"{minutes}m {seconds:.3f}s"


class CustomLogger(logging.Logger):
    """
    Custom logger class that uses the CustomRelativeFormatter.
    
    Attributes:
        level (int): The logging level for the logger.

    format:
        '[%(asctime)s]: %(message)s'
    datefmt:
        '%H:%M:%S'
    """
    
    def __init__(self, level=logging.INFO):
        super().__init__(level)
        self.setLevel(level)
        handler = logging.StreamHandler()
        formatter = CustomRelativeFormatter('[%(asctime)s]: %(message)s')
        handler.setFormatter(formatter)
        self.addHandler(handler)
        print(time.strftime("[%d/%m %H:%M:%S]", time.localtime()), ": Logger initialized")

def get_api_key(file_path='.api_key') -> str:
    try:
        with open(file_path, 'r') as f:
            api_key = f.read().strip()
        return api_key
    except FileNotFoundError:
        print(f"API key file not found: {file_path}")
        raise


def create_time_hash() -> str:
    """
    Create a time-based hash string.
    
    Returns:
        str: A string representing the current time in the format 'YYYYMMDD_HHMMSS'.
    """
    h = sha1()
    h.update(datetime.now().strftime("%Y%m%d%H%M%S%f").encode('utf-8'))
    time_hash = h.hexdigest()

    return time_hash


def retrieve_model_info(model_name: str = None, csv_file: str = "models.csv") -> Tuple[str, str]:
    """
    Retrieve model digest number and kv cache size from the model instance.

    Args:
        model (str): The name of the model to retrieve information from. Defaults to None.
        csv_file (str): The path to the CSV file containing model information. Defaults to "models.csv".
    Returns:
        Tuple[str, str]: A tuple containing the model digest number and kv cache size.
    """
    if not isinstance(model_name, str):
        raise TypeError("model_name must be a string")
    if not os.path.exists(csv_file):
        raise FileNotFoundError(f"CSV file not found: {csv_file}")


    models_DF = pd.read_csv(csv_file)
    
    match = models_DF[models_DF['model_name'] == model_name]

    if not match.empty:
        digest = match.iloc[0]['digest_nr']
        kv_cache = match.iloc[0]['kv_cache']
        return digest, kv_cache
    else:
        raise NameError(f"Did not find model: {model_name}")
    

def write_response_to_csv(model_name: str,
                          prompt_id: int, 
                          prompt: str, 
                          response: str, 
                          file_name: str,
                          time_hash: str,
                          append: bool = True,
                          prompt_instructions: str = None
                          ) -> None:
    """
        Write a prompt and its response to a CSV file.
    Args:
        prompt (str): The prompt to write to the CSV file.
        response (str): The response to write to the CSV file.
    Returns:
        None
    """

    if not isinstance(prompt, str):
        raise TypeError("prompt must be a string")
    if not isinstance(response, str):
        raise TypeError("response must be a string")
    if not file_name.endswith('.csv'):
        raise ValueError("file_name must be a .csv file")
    if not os.path.exists(file_name):
        # Create a new CSV file with headers if it does not exist
        df_header = pd.DataFrame({
            "model": [],
            "prompt id": [],
            "prompt": [],
            "response": [],
            "hash": []
        })
        if prompt_instructions:
            df_header["prompt instructions"] = []
        df_header.to_csv(file_name, index=False)

    # Append the prompt and response to the CSV file
    df = pd.DataFrame({
        "model": [model_name],
        "prompt id": [prompt_id],
        "prompt": [prompt],
        "response": [response],
        "hash": [time_hash]
    })
    if prompt_instructions:
        df["prompt instructions"] = prompt_instructions
    if append:
        df.to_csv(file_name, mode='a', header=False, index=False)
    else:
        df.to_csv(file_name, mode='w', header=True, index=False)


def read_responses_from_csv(file_name: str) -> pd.DataFrame:
    """
    Read responses from a CSV file and return them as a DataFrame.
    
    Args:
        file_name (str): The name of the CSV file to read from. Defaults to "responses.csv".
    Returns:
        pd.DataFrame: A DataFrame containing the prompts and their corresponding responses.
    """
    if not os.path.exists(file_name):
        raise FileNotFoundError(f"CSV file not found: {file_name}")
    
    df = pd.read_csv(file_name, sep=',', encoding='utf-8', encoding_errors='ignore')
    return df


def write_to_xlsx(df: pd.DataFrame, file_name: str, sheet_name: str, test_type: TypeOfTest) -> None:
    """
    Write data to an Microsoft Excel file (.xlsx), appending to a specified sheet.
    If the sheet does not exist, it will be created.
    If the file does not exist, it will be created.
    
    Args:
        new_data (pd.DataFrame): The data to write to the Excel file.
        file_name (str): The name of the Excel file.
        sheet (str): The name of the sheet to write to.
    """
    if not Path(file_name).exists():
        initNewExcel(test_type=test_type, fileName=file_name)
        print('init')
    workbook = load_workbook(file_name)

    header = False
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        startrow = sheet.max_row
    else:
        sheet = workbook.create_sheet(sheet_name)
        startrow = 0
        header = True

    # Use ExcelWriter in append mode without setting writer.book
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=header, startrow=startrow)

def initNewExcel(test_type: TypeOfTest, fileName: str):
    """
    Initiates blank excel with headers
    """
    match test_type:
        case TypeOfTest.SUMMARIZATION | TypeOfTest.PROMPT_ALIGNMENT | TypeOfTest.HELPFULNESS:
            sheet_map = {
                TypeOfTest.SUMMARIZATION: "SummarizationResults",
                TypeOfTest.PROMPT_ALIGNMENT: "PromptAlignmentResults",
                TypeOfTest.HELPFULNESS: "HelpfulnessResults"
            }

            df_header = pd.DataFrame({
                "Model": [],
                "Digest": [],
                "KV Cache": [],
                "Prompt Number": [],
                "Judge Model": [],
                "Judge Seed": [],
                "Judge Temperature": [],
                "Score": [],
                "Reason": [],
                "Hash": []
            })

            with pd.ExcelWriter(filepath) as writer:
                df_header.to_excel(writer, index=False, sheet_name=sheet_map[test_type])

        case TypeOfTest.BENCHMARKING:
            df = pd.DataFrame({
                'Model': [],
                'Digest': [],
                'KV Cache Type': [],
                'Prompt nr':[],
                'Total Duration[ms]': [],
                'Load Duration[ms]':[],
                'Promt Eval Count':[],
                'Prompt eval duration[ms]':[],
                'Prompt eval rate':[],
                'Eval Count':[],
                'Eval duration[ms]':[],
                'Eval rate':[],
                'Inteded purpose': []
            })

            avg_df = pd.DataFrame({
                'Model':[],
                'Digest': [],
                'KV Cache Type': [],
                'Average time[s]': [],
                'Average tokens/s':[],
                'Inteded purpose': [],
                'Language errors': []
            })

            # Create excel outside of git repo
            filepath = Path.cwd()/fileName

            with pd.ExcelWriter(filepath) as writer:
                df.to_excel(writer, index=False, sheet_name='Benchmarks')
                avg_df.to_excel(writer, index=False, sheet_name='Avg_Benchmarks')
        case _:
            raise ValueError("Invalid test type provided. Use TestType.SUMMARIZATION, TestType.PROMPT_ALIGNMENT, TestType.BENCHMARKING or TestType.HELPFULNESS.")



    
def save_eval_results_to_xlsx(
        type_of_test: TypeOfTest,
        model_name: str,
        results: List[tuple],
        file_name: str,
        judge_params: Tuple[str, int, float, int],
        prompt_id: int = None,
        time_hash: str = ""
        ) -> None:
    """Log the summarization results to .xlsx file.
    Args:
        type_of_test (TestType): The type of test being evaluated.
        model_name (str): The name of the model used for summarization.
        results (List[tuple]): A list of tuples containing the score and reason for each prompt.
        file_name (str): The name of the file to save the results.
        judge_params (Tuple[str, int, float, int]): A tuple containing the judge model parameters:
            - str: The name of the judge model
            - int: The seed for the judge model
            - float: The temperature for the judge model
            - int: The top_k for the judge model
        prompt_id (int, optional): The ID of the prompt being evaluated. Defaults to None.
        time_hash (str, optional): A hash string representing the time of evaluation. Defaults to "".
    Raises:
        ValueError: If the file name does not end with '.xlsx'.
        TypeError: If type_of_test is not an instance of TestType.
        ValueError: If type_of_test is not one of the defined TestType values.
    Returns:
        None: This function does not return anything. It writes the results to an Excel file.
    """
    if not file_name.endswith('.xlsx'):
        raise ValueError("File name is not an .xlsx file")
    if not isinstance(type_of_test, TypeOfTest):
        raise TypeError("type_of_test must be a TestType enum")
    
    match type_of_test:
        case TypeOfTest.SUMMARIZATION:
            sheet_name = "SummarizationResults"
        case TypeOfTest.PROMPT_ALIGNMENT:
            sheet_name = "PromptAlignmentResults"
        case TypeOfTest.HELPFULNESS:
            sheet_name = "HelpfulnessResults"
        case _:
            raise ValueError("Invalid test type provided. Use TestType.SUMMARIZATION, TestType.PROMPT_ALIGNMENT, or TestType.HELPFULNESS.")

    digest, kv_cache = retrieve_model_info(model_name=model_name)

    judge_model_name, judge_seed, judge_temp, judge_top_k = judge_params


    for prompt_no, (score, reason) in enumerate(results):
        if prompt_id is not None:
            prompt_ref = prompt_id
        else:
            prompt_ref = prompt_no

        df_result = pd.DataFrame({
            "Model": [model_name],
            "Digest": [digest],
            "KV Cache": [kv_cache],
            "Prompt Number": [prompt_ref],
            "Judge Model": [judge_model_name],
            "Judge Seed": [judge_seed],
            "Judge Temp": [judge_temp],
            "Judge Top K": [judge_top_k],
            "Score": [score],
            "Reason": [reason],
            "Hash": [time_hash]
        })

        write_to_xlsx(
            df=df_result, 
            file_name=file_name, 
            sheet_name=sheet_name,
            test_type = type_of_test
        )
            
